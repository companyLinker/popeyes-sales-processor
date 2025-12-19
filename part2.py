# part2.py
import os
import io
import re
import pandas as pd
from datetime import datetime, timedelta
import dateutil.parser
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
import datetime as dt_module

# ================= CONFIGURATION =================
SOURCE_ROOT_ID = "16edTsOusrYf-5LqRgiGqIwMn94H6yzsE"     # Converted CSVs
DEST_ROOT_ID = "1tlPuBOhnxjQJ_kIGo7-WW6TjG2mxfbgr"       # Final Excel files
TRACKING_SHEET_ID = "1r872UNCcsgkdEkV9Y9PnNcuTtPrezs0XE3n8HFZgqyM"       # <<< SAME AS PART1

SERVICE_ACCOUNT_JSON = json.loads(os.environ['SERVICE_ACCOUNT_KEY'])

SCOPES = ['https://www.googleapis.com/auth/drive',
          'https://www.googleapis.com/auth/spreadsheets']

creds = Credentials.from_service_account_info(SERVICE_ACCOUNT_JSON, scopes=SCOPES)
drive_service = build('drive', 'v3', credentials=creds)
sheets_service = build('sheets', 'v4', credentials=creds)

# ================= TRACKING HELPERS =================
def get_part1_done_ids():
    try:
        result = sheets_service.spreadsheets().values().get(
            spreadsheetId=TRACKING_SHEET_ID, range="Sheet1!A:D").execute()
        rows = result.get('values', [])
        if len(rows) <= 1: return set()
        return {row[0] for row in rows[1:] if len(row) >= 4 and row[3] == "PART1_DONE"}
    except: return set()

def get_part2_done_ids():
    try:
        result = sheets_service.spreadsheets().values().get(
            spreadsheetId=TRACKING_SHEET_ID, range="Sheet1!A:D").execute()
        rows = result.get('values', [])
        if len(rows) <= 1: return set()
        return {row[0] for row in rows[1:] if len(row) >= 4 and row[3] == "PART2_DONE"}
    except: return set()

def log_to_sheet(file_id, file_name, stage):
    timestamp = dt_module.datetime.now().isoformat()
    body = {'values': [[file_id, file_name, timestamp, stage]]}
    sheets_service.spreadsheets().values().append(
        spreadsheetId=TRACKING_SHEET_ID,
        range="Sheet1!A:D",
        valueInputOption="RAW",
        body=body
    ).execute()

# ================= DRIVE HELPERS =================
def get_or_create_folder(parent_id, folder_name):
    query = f"'{parent_id}' in parents and mimeType='application/vnd.google-apps.folder' and name='{folder_name}' and trashed=false"
    results = drive_service.files().list(q=query, fields="files(id)").execute()
    files = results.get('files', [])
    if files:
        return files[0]['id']
    metadata = {'name': folder_name, 'mimeType': 'application/vnd.google-apps.folder', 'parents': [parent_id]}
    folder = drive_service.files().create(body=metadata, fields='id').execute()
    print(f"Created folder: {folder_name}")
    return folder['id']

def download_csv_to_df(file_id):
    request = drive_service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    fh.seek(0)
    return pd.read_csv(fh, dtype=str, low_memory=False)

def get_date_file_logic(dt):
    if pd.isna(dt):
        return pd.NA
    time = dt.time()
    if time >= datetime.strptime('03:00:00', '%H:%M:%S').time():
        return dt.strftime('%m/%d/%Y').lower()
    else:
        return (dt - timedelta(days=1)).strftime('%m/%d/%Y').lower()

# ================= STORE PROCESSING (Your Full Original Logic) =================
def process_store_batch(store_name, new_files_for_store, dest_store_folder_id):
    print(f"Processing store: {store_name} ({len(new_files_for_store)} new files)")

    try:
        df_list = []
        for file_item in new_files_for_store:
            df_temp = download_csv_to_df(file_item['id'])
            df_temp['filename'] = os.path.splitext(file_item['name'])[0]
            df_temp.columns = df_temp.columns.str.strip()

            if 'Date_time' in df_temp.columns:
                df_temp['Date_time'] = df_temp['Date_time'].str.replace(',', '', regex=False)
                df_temp['Date_time'] = df_temp['Date_time'].apply(lambda x: dateutil.parser.parse(x) if pd.notnull(x) else pd.NaT)
                df_temp['Date_file'] = df_temp['Date_time'].apply(get_date_file_logic)
                df_temp['display_date'] = (df_temp['Date_time'] - pd.Timedelta(minutes=1)).dt.strftime('%m/%d/%Y %I:%M%p').str.upper()
                df_temp.insert(0, 'Date_file', df_temp.pop('Date_file'))
                df_list.append(df_temp)

        if not df_list:
            return

        df_new = pd.concat(df_list, ignore_index=True)

        # Find split columns
        split_0_col = [c for c in df_new.columns if c.endswith('_split_0')][0]
        split_1_col = [c for c in df_new.columns if c.endswith('_split_1')][0]
        split_3_col = [c for c in df_new.columns if c.endswith('_split_3')][0]
        split_5_col = [c for c in df_new.columns if c.endswith('_split_5')][0]

        prefix = split_0_col.replace('_split_0', '')
        split_35_col = prefix + '_split_35'

        df_new[split_5_col] = pd.to_numeric(df_new[split_5_col], errors='coerce')
        df_new[split_3_col] = pd.to_numeric(df_new[split_3_col], errors='coerce')
        df_new[split_35_col] = df_new[split_5_col] * df_new[split_3_col]

        # Output file
        output_filename = f"{store_name}_Consolidated_data.xlsx".replace(" ", "_")
        local_path = f"/tmp/{output_filename}"

        # Load existing or start fresh
        existing = drive_service.files().list(
            q=f"'{dest_store_folder_id}' in parents and name='{output_filename}' and trashed=false",
            fields="files(id)"
        ).execute().get('files', [])

        if existing:
            print("Appending to existing consolidated file")
            existing_id = existing[0]['id']
            req = drive_service.files().get_media(fileId=existing_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, req)
            done = False
            while not done:
                status, done = downloader.next_chunk()
            fh.seek(0)
            df_existing = pd.read_excel(fh, sheet_name='Data')
            df_full = pd.concat([df_existing, df_new], ignore_index=True)
        else:
            print("Creating new consolidated file")
            df_full = df_new

        # Save full data
        df_full.to_excel(local_path, sheet_name='Data', index=False)

        # ================= YOUR FULL PIVOT LOGIC (EXACTLY AS BEFORE) =================
        categories = [
            '10000000,', '30000000,', '30004001,', '30004002,', '30004003,', '30004004,', '30006007,', '30004029,', '30009100,',
            '30009101,', '30009102,', '30009103,', '30009112,', '30009113,', '30009114,', '30009115,', '30009131,',
            '40001001,', '40001002,', '40001003,', '40002002,', '7019900,', '40001004', '30009123,', '30009120,', '30009122,', '30009121,', '30009129,',
            '30009092,', '30009093,', '30009094,', '30009095,', '30009096,', '30009097,', '30009098,', '30009099,',
            '30009100,', '30009101,', '30009102,', '30009103,', '30009104,', '30009105,', '30009106,', '30009107,',
            '30009108,', '30009109,', '30009110,', '30009111,', '30009112,', '30009113,', '30009114,', '30009115,',
            '30009131,', '30009132,', '30009133,', '30009134,', '30009135,', '30009136,', '30004007,', '40002010,',
            '19999984,', '19999980,', '7019395,', '40002001,', '9001600,', '30003010,', '40002011,', '7019910,',
            '30009145,', '30009146,', '30009147,', '30009148,', '30009149,', '30009150,', '30009151,', '30009152,',
            '30009153,', '30009154,', '30009155,', '30006006,', '30009124,', '30009125,', '30009126,', '30009129,',
            '30009127,', '30004055,', '30004035,', '30004035,'
        ]
        categories2 = [
            '30004025,', '30004024,', '30004026,', '30004027,', '20000033,', '20000030,', '20000031,',
            '19999999,', '20000000,', '20000005,', '20000006,', '20000010,', '20000011,', '20000015,', '30009112,',
            '30009113,', '30009114,', '30009115,', '30009122,', '30009123,', '30009146,', '30009149,', '30009151,', '30009154,'
        ]
        categories3_bev = [
            '20000002,', '29000160,', '80101,', '80102,', '80103,', '80201,',
            '80202,', '80203,', '80301,', '80302,', '80303,', '80601,', '80602,', '80603,'
        ]
        donation_key = ['7019910,']

        cc = set(categories) | set(categories2) | set(categories3_bev)
        cc1 = list(cc)
        cc_bev = set(categories2) | set(categories3_bev)
        cc2 = list(cc_bev)
        ccd = set(categories) | set(categories2)
        ccd1 = list(ccd)
        don = set(donation_key)
        dona = list(don)

        category_filter1 = df_full[split_0_col].isin(cc1)
        category_filter2 = df_full[split_0_col].isin(categories2)
        category_filter4 = df_full[split_0_col].isin(ccd1)
        category_filter5 = df_full[split_0_col].isin(dona)

        filtered_df21 = df_full[category_filter1].copy()
        pivot_table11 = filtered_df21.pivot_table(index=['Date_time', 'Date_file'], columns=[split_0_col, split_1_col], values=split_35_col, aggfunc="sum")

        filtered_df22 = df_full[category_filter2].copy()
        pivot_table22 = filtered_df22.pivot_table(index=['Date_time', 'Date_file'], columns=[split_0_col, split_1_col], values=split_5_col, aggfunc="sum")

        filtered_df33 = df_full[category_filter4].copy()
        pivot_table33 = filtered_df33.pivot_table(index=['Date_time', 'Date_file'], columns=[split_0_col, split_1_col], values=[split_5_col, split_35_col], aggfunc="sum")

        filtered_df34 = df_full[category_filter5].copy()
        pivot_table34 = filtered_df34.pivot_table(index=['Date_time', 'Date_file'], columns=[split_0_col, split_1_col], values=[split_5_col], aggfunc="sum")

        # Add totals
        pivot_table11_with_totals = pivot_table11.groupby(level=1, observed=True).apply(lambda x: x._append(x.sum().rename((x.name, 'Total'))))
        pivot_table12_with_totals = pivot_table11.groupby(level=1, observed=True).apply(lambda x: x.sum().rename((x.name, 'Total')))
        pivot_table22_with_totals = pivot_table22.groupby(level=1, observed=True).apply(lambda x: x._append(x.sum().rename((x.name, 'Total'))))
        pivot_table33_with_totals = pivot_table33.groupby(level=1, observed=True).apply(lambda x: x._append(x.sum().rename((x.name, 'Total'))))
        pivot_table34_with_totals = pivot_table34.groupby(level=1, observed=True).apply(lambda x: x.sum().rename((x.name, 'Total')))

        # Write all sheets
        with pd.ExcelWriter(local_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            pivot_table22_with_totals.to_excel(writer, sheet_name='Pivot_Delv')
            pivot_table11_with_totals.to_excel(writer, sheet_name='PivotTable_total')
            pivot_table12_with_totals.to_excel(writer, sheet_name='Total_summary')
            pivot_table34_with_totals.to_excel(writer, sheet_name='Donation')
            pivot_table33_with_totals.to_excel(writer, sheet_name='Soda_dinein_sales')

        # Customer Count
        wb = load_workbook(local_path)
        df_pivot_total = pd.read_excel(local_path, sheet_name='PivotTable_total', header=2)
        df_pivot_total['Date_time'] = pd.to_datetime(df_pivot_total['Date_time'], errors='coerce')
        df_pivot_total['Date_only'] = df_pivot_total['Date_time'].apply(get_date_file_logic)

        customer_count_df = df_pivot_total.groupby('Date_only').size().reset_index(name='Customer_Count')
        pivot_table_cnt = pd.pivot_table(df_pivot_total, values='Date_time', index='Date_only', aggfunc='count').reset_index()
        pivot_table_cnt.columns = ['Date_only', 'Total Count']
        result_df = pd.merge(customer_count_df, pivot_table_cnt, on='Date_only', how='left')

        with pd.ExcelWriter(local_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            result_df.to_excel(writer, index=False, sheet_name='Customer_Count')

        # Formatting
        wb = load_workbook(local_path)
        for ws in wb.worksheets:
            last_row = ws.max_row
            last_col = ws.max_column
            ws.auto_filter.ref = f"A3:{get_column_letter(last_col)}{last_row}"
            ws.freeze_panes = "D4"
        wb.save(local_path)

        # Upload / Update
        media = MediaFileUpload(local_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        if existing:
            drive_service.files().update(fileId=existing_id, media_body=media).execute()
            print(f"Updated {output_filename}")
        else:
            metadata = {'name': output_filename, 'parents': [dest_store_folder_id]}
            drive_service.files().create(body=metadata, media_body=media).execute()
            print(f"Created {output_filename}")

        os.remove(local_path)

        # Mark files as done
        for f in new_files_for_store:
            log_to_sheet(f['id'], f['name'], "PART2_DONE")

    except Exception as e:
        print(f"Error processing {store_name}: {e}")

# ================= MAIN – Group New Converted Files by Store =================
def main():
    part1_done = get_part1_done_ids()
    part2_done = get_part2_done_ids()
    new_converted = []

    def recursive_collect(folder_id):
        page_token = None
        while True:
            results = drive_service.files().list(
                q=f"'{folder_id}' in parents and trashed=false",
                fields="nextPageToken, files(id, name, mimeType, parents)",
                pageToken=page_token
            ).execute()
            items = results.get('files', [])
            for item in items:
                if item.get('mimeType') == 'application/vnd.google-apps.folder':
                    recursive_collect(item['id'])
                elif item['name'].startswith('converted_') and item['name'].endswith('.csv'):
                    if item['id'] in part1_done and item['id'] not in part2_done:
                        new_converted.append(item)
            page_token = results.get('nextPageToken')
            if not page_token:
                break

    print("Scanning for new converted files...")
    recursive_collect(SOURCE_ROOT_ID)
    print(f"Found {len(new_converted)} file(s) ready for PART2.")

    # Group by store (parent folder name)
    store_groups = {}
    for file_item in new_converted:
        parents = drive_service.files().get(fileId=file_item['id'], fields="parents").execute().get('parents', [])
        if parents:
            parent = drive_service.files().get(fileId=parents[0], fields="name").execute()
            store_name = parent['name']
            store_groups.setdefault(store_name, []).append(file_item)

    processed_stores = 0
    for store_name, files in store_groups.items():
        dest_id = get_or_create_folder(DEST_ROOT_ID, store_name)
        process_store_batch(store_name, files, dest_id)
        processed_stores += 1

    print(f"\nPART2 Complete – Updated {processed_stores} store(s).")

if __name__ == "__main__":
    main()